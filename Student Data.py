from tkinter import *
from tkinter import ttk
from tkinter.ttk import Combobox
from tkinter  import messagebox
import openpyxl
import xlrd
from openpyxl import Workbook
import pathlib

#window properties
root = Tk() #initialize
root.title("Excel Data Form") #title
root.geometry("700x400+300+200") #dimensions
root.resizable(0, 0) #resizability
root.configure(bg='#326273') #background
#window icon
icon_image = PhotoImage(file='icon.png')
root.iconphoto(False, icon_image)

#app logo
logo = PhotoImage(file='logo.png')
Label(root, image=logo, bg='#326273').place(x=600, y=5)

file=pathlib.Path('FormDB.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] = "Firstname"
    sheet['B1'] = "Middlename"
    sheet['C1'] = "Surname"
    sheet['D1'] = "Phone Number"
    sheet['E1'] = "Email"
    sheet['F1'] = "Age"
    sheet['G1'] = "Gender"

    file.save('FormDB.xlsx')

def submit():
    firstname = f_nameValue.get()
    middlename = m_nameValue.get()
    surname = s_nameValue.get()
    phone = phoneValue.get()
    email = emailValue.get()
    age = ageValue.get()
    gender = genderValue.get()

    # print(firstname)
    # print(middlename)
    # print(surname)
    # print(phone)
    # print(email)
    # print(age)
    # print(gender)

    file = openpyxl.load_workbook('FormDB.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value = firstname)
    sheet.cell(column=2, row=sheet.max_row, value = middlename)
    sheet.cell(column=3, row=sheet.max_row, value = surname)
    sheet.cell(column=4,row=sheet.max_row,  value = phone)
    sheet.cell(column=5,row=sheet.max_row,  value = email)
    sheet.cell(column=6, row=sheet.max_row, value = age)
    sheet.cell(column=7, row=sheet.max_row, value = gender)

    file.save(r'FormDB.xlsx')
    messagebox.showinfo('info', 'New entry created')

    f_nameValue.set("")
    m_nameValue.set("")
    s_nameValue.set("")
    phoneValue.set("")
    emailValue.set("")
    ageValue.set("")
    genderValue.set("")

def clear():
    #clear all fields
    f_nameValue.set("")
    m_nameValue.set("")
    s_nameValue.set("")
    phoneValue.set("")
    emailValue.set("")
    ageValue.set("")
    genderValue.set("")
    
    
    

#heading
heading = Label(root, text="Please fill out the form below", font="Verdana 14", bg="#326273", fg="#fff")
heading.place(x=20, y=20)

#labels
f_name_label = Label(root, text="Firstname", font="Verdana 11", bg='#326273', fg='#fff')
f_name_label.place(x=50, y=100)
m_name_label = Label(root, text="Middlename", font="Verdana 11", bg='#326273', fg='#fff')
m_name_label.place(x=50, y=150)
s_name_label = Label(root, text="Surname", font="Verdana 11", bg='#326273', fg='#fff')
s_name_label.place(x=50, y=200)
phone_label = Label(root, text="Phone", font="Verdana 11", bg='#326273', fg='#fff')
phone_label.place(x=50, y=250)
email_label = Label(root, text="Email", font="Verdana 11", bg='#326273', fg='#fff')
email_label.place(x=350, y=250)
age_label = Label(root, text="Age", font="Verdana 11", bg='#326273', fg='#fff')
age_label.place(x=50, y=300)
gender_label = Label(root, text="Gender", font="Verdana 11", bg='#326273', fg='#fff')
gender_label.place(x=350, y=300)


#entry
f_nameValue = StringVar()
m_nameValue = StringVar()
s_nameValue = StringVar()
phoneValue = StringVar()
emailValue = StringVar()
ageValue = StringVar()
genderValue = StringVar()


f_nameEntry = Entry(root, textvariable = f_nameValue, width=50, bd=2, font=20)
f_nameEntry.place(x=150, y=100)
m_nameEntry = Entry(root, textvariable = m_nameValue, width=50, bd=2, font=20)
m_nameEntry.place(x=150, y=150)
s_nameEntry = Entry(root, textvariable = s_nameValue, width=50, bd=2, font=20)
s_nameEntry.place(x=150, y=200)
pnoneEntry = Entry(root, textvariable = phoneValue, width=20, bd=2, font=20)
pnoneEntry.place(x=150, y=250)
emailEntry = Entry(root, textvariable = emailValue, width=20, bd=2, font=20)
emailEntry.place(x=420, y=250)
ageEntry = Entry(root, textvariable = ageValue, width=20, bd=2, font=20)
ageEntry.place(x=150, y=300)
genderCombo = Combobox(root, values= ['Male', 'Female', 'Please Select'], font='Verdana 12', state='r', width=17)
genderCombo.place(x=420, y=300)
genderCombo.set("Please Select")

#submit button
submit_button = Button(root, text='Submit', bg= '#326273', fg='#fff', font='Verdana 11', width= 10, height=0, command=submit)
# , ,
submit_button.place(x=50, y=350)
clear_button = Button(root, text='Clear', bg= '#326273', fg='#fff', font='Verdana 11', width= 10, height=0, command=clear)
clear_button.place(x=300, y=350)
exit_button = Button(root, text='Exit', bg= '#326273', fg='#fff', font='Verdana 11', width= 10, height=0, command=lambda:root.destroy())
exit_button.place(x=550, y=350)


root.mainloop()